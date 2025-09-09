import winreg
import platform
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def get_installed_software():
    uninstall_keys = [
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
        r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
    ]

    software_list = []

    for uninstall_key in uninstall_keys:
        try:
            reg_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, uninstall_key)
        except FileNotFoundError:
            continue

        for i in range(0, winreg.QueryInfoKey(reg_key)[0]):
            try:
                sub_key_name = winreg.EnumKey(reg_key, i)
                sub_key = winreg.OpenKey(reg_key, sub_key_name)
                software = {}

                try:
                    software["Name"] = winreg.QueryValueEx(sub_key, "DisplayName")[0]
                except FileNotFoundError:
                    continue

                try:
                    software["Version"] = winreg.QueryValueEx(sub_key, "DisplayVersion")[0]
                except FileNotFoundError:
                    software["Version"] = "unknown"

                try:
                    software["Publisher"] = winreg.QueryValueEx(sub_key, "Publisher")[0]
                except FileNotFoundError:
                    software["Publisher"] = "unknown"

                try:
                    software["language"] = winreg.QueryValueEx(sub_key, "Language")[0]
                except FileNotFoundError:
                    software["language"] = "unknown"

                try:
                    software["Product code"] = winreg.QueryValueEx(sub_key, "ProductID")[0]
                except FileNotFoundError:
                    software["Product code"] = "unknown"

                if "WOW6432Node" in uninstall_key:
                    software["Bits"] = "32"
                else:
                    software["Bits"] = "64" if platform.machine().endswith("64") else "32"

                software["kind"] = "unknown"

                software_list.append(software)

            except Exception:
                continue

    return software_list


def export_to_excel(software_list, filename="software list.xlsx", Name=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Programmes"

    # styles
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")

    # Row 2
    ws.merge_cells("B2:F2")
    ws["B2"] = "Programmes on my laptop"
    ws["B2"].font = Font(bold=True)
    ws["B2"].alignment = center_align

    ws["H2"] = f"Jouw Name: {Name}"
    ws["H2"].alignment = center_align

    # Row 3
    headers = ["Name van de software", "Version", "language", "Bits (32/64)", "Publisher", "Product code", "kind programma"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Data from row 4+
    for row_idx, sw in enumerate(software_list, start=4):
        ws.cell(row=row_idx, column=2, value=sw["Name"])
        ws.cell(row=row_idx, column=3, value=sw["Version"])
        ws.cell(row=row_idx, column=4, value=sw["language"])
        ws.cell(row=row_idx, column=5, value=sw["Bits"])
        ws.cell(row=row_idx, column=6, value=sw["Publisher"])
        ws.cell(row=row_idx, column=7, value=sw["Product code"])
        ws.cell(row=row_idx, column=8, value=sw["kind"])

    wb.save(filename)
    print(f"Excel file saved as: {filename}")


if __name__ == "__main__":
    programs = get_installed_software()
    export_to_excel(programs, "software list.xlsx", Name="[name]")
